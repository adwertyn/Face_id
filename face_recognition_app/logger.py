import os
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def log_attendance(name, class_name):
    folder = "attendance_logs"
    os.makedirs(folder, exist_ok=True)
    date = datetime.now().strftime("%Y-%m-%d")

    file_path = os.path.join(folder, f"{class_name}_{date}.txt")
    all_students_path = os.path.join("known_faces", class_name, "students.txt")

    if os.path.exists(file_path):
        with open(file_path, "r", encoding="utf-8") as f:
            present_students = {line.strip() for line in f.readlines()}
    else:
        present_students = set()

    present_students.add(name)

    # Save updated attendance
    with open(file_path, "w", encoding="utf-8") as f:
        for student in sorted(present_students):
            f.write(student + "\n")

def get_attendance_text(class_name):
    folder = "attendance_logs"
    date = datetime.now().strftime("%Y-%m-%d")
    attendance_file = os.path.join(folder, f"{class_name}_{date}.txt")
    all_students_file = os.path.join("known_faces", class_name, "students.txt")

    if not os.path.exists(all_students_file):
        return "⚠️ students.txt not found for this class."

    with open(all_students_file, "r", encoding="utf-8") as f:
        all_students = [line.strip() for line in f.readlines() if line.strip()]

    present_students = set()
    if os.path.exists(attendance_file):
        with open(attendance_file, "r", encoding="utf-8") as f:
            present_students = {line.strip() for line in f.readlines()}

    report_lines = [f"📋 Attendance for {class_name} – {date}\n"]
    for student in all_students:
        mark = "✅" if student in present_students else "❌"
        report_lines.append(f"{mark} {student}")

    return "\n".join(report_lines)

def export_attendance_excel(class_name):
    folder = "attendance_logs"
    detailed_folder = "excel_reports"
    os.makedirs(detailed_folder, exist_ok=True)
    date = datetime.now().strftime("%Y-%m-%d")
    time_now = datetime.now().strftime("%H:%M:%S")

    attendance_file = os.path.join(folder, f"{class_name}_{date}.txt")
    all_students_file = os.path.join("known_faces", class_name, "students.txt")
    excel_path = os.path.join(detailed_folder, f"{class_name}_attendance_{date}.xlsx")

    if not os.path.exists(all_students_file):
        return None

    with open(all_students_file, "r", encoding="utf-8") as f:
        all_students = [line.strip() for line in f.readlines() if line.strip()]

    present_students = {}
    if os.path.exists(attendance_file):
        with open(attendance_file, "r", encoding="utf-8") as f:
            for line in f:
                name = line.strip()
                present_students[name] = f"{date} {time_now}"

    df = pd.DataFrame(columns=["Student Name", "Status", "Time"])

    for student in all_students:
        if student in present_students:
            df = df._append({
                "Student Name": student,
                "Status": "✅ Present",
                "Time": present_students[student]
            }, ignore_index=True)
        else:
            df = df._append({
                "Student Name": student,
                "Status": "❌ Absent",
                "Time": "Kelmagan"
            }, ignore_index=True)

    df.to_excel(excel_path, index=False)

    wb = load_workbook(excel_path)
    ws = wb.active

    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for row in range(2, len(df) + 2):
        status = ws[f"B{row}"].value
        if "Present" in status:
            ws[f"B{row}"].fill = green_fill
            ws[f"C{row}"].fill = green_fill
        else:
            ws[f"B{row}"].fill = red_fill
            ws[f"C{row}"].fill = red_fill

    wb.save(excel_path)
    return excel_path
